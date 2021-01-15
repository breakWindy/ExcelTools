#######
# here is tool of checking excel , which is check files full same.
#######

from openpyxl import load_workbook

class excelCell:
    def __init__(self, sheetName, rowNumber, columnNumber, strCellContent):
        self.sheetName = sheetName
        self.rowNumber = rowNumber
        self.columnNumber = columnNumber
        self.strCellContent = strCellContent

    def isSameCellContent(self, originCell):
        return self.strCellContent == originCell.strCellContent

    def __str__(self):
        return '\n sheetName={}, rowNumber={}, column={} , strCellContent={} \n'.format(self.sheetName, self.rowNumber, self.columnNumber, self.strCellContent)


def extractAllExcelCell(excelPath, checkSheetName):

    workbook = load_workbook(excelPath, data_only=True)
    excelCellList = []

    for n , sheet in enumerate(workbook.worksheets):

        if sheet.title == checkSheetName:
            for rowNumber, row_cells in enumerate(sheet.iter_rows()):
                for columnNumber, cell in enumerate(row_cells):
                # here is create cell of excelCell ,  than collect all , compare all checkExcelCell
                    cellValue = cell.value
                    if cell.data_type == 'n' and cell.value != None:
                        cellValue= round(float(cell.value), 2)

                    xcell = excelCell(sheet.title, cell.row, cell.column, cellValue)
                    excelCellList.append(xcell)

    return excelCellList

def output(content):
    print(content)

def outputOC(originContent, checkContent):
    output('originCell: ' + originContent)
    output('checkCell: ' + checkContent)


def checkCellList(originCellList, checkCellList):

    for n , cell in enumerate(originCellList):
        if n >= len(checkCellList):
            outputOC(str(cell), 'not exist cell')
            continue

        if not cell.isSameCellContent(checkCellList[n]):
            outputOC(str(cell), str(checkCellList[n]))


def __main__():
    originExcelPath = input('Please input originExcelPath: ')
    checkExcelPath = input('Please input checkExcelPath: ')

    checkSheetName=input('Please input name of checked sheet: ')

#    originExcelPath = '/Users/luowen/Downloads/origin.xlsx'
#    checkExcelPath = '/Users/luowen/Downloads/check.xlsx'
#    checkSheetName = '余额表简表'

#    originExcelPath = '/Users/luowen/Downloads/JTZ-20201231现金流量表C.xlsx'
#    checkExcelPath = '/Users/luowen/Downloads/JTZ-20201231基础7张表_新现金流量表.xlsx'
#    checkSheetName = '现金流量表新表'

    originExcelCellList = extractAllExcelCell(originExcelPath, checkSheetName)
    checkExcelCellList = extractAllExcelCell(checkExcelPath, checkSheetName)

    checkCellList(originExcelCellList, checkExcelCellList)

    output('originCellSize={}, checkCellSize={} '.format(len(originExcelCellList), len(checkExcelCellList)))

__main__()
